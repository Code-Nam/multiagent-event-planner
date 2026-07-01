"""
create_templates.py — AGEVP

Creates blank starter template files in templates/:
  templates/recap.xlsx         — Excel workbook template
  templates/report.docx        — Word document template
  templates/presentation.pptx  — PowerPoint template

Customize these files (colors, fonts, logos) in LibreOffice / Word /
PowerPoint, then re-save. The generation scripts will pick them up
automatically on the next run.

Usage:
    python scripts/create_templates.py
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from docx import Document
    from docx.shared import Pt, RGBColor
    from pptx import Presentation
    from pptx.util import Pt as PPTXPt
except ImportError:
    print("Missing dependency. Run: pip install -r scripts/requirements.txt", file=sys.stderr)
    sys.exit(1)

TEMPLATES_DIR = Path(__file__).parent.parent / "templates"


def create_xlsx_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    # Example branded header row — edit in LibreOffice/Excel to set real colors
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col, value=f"Header {col}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 20
    wb.save(path)


def create_docx_template(path: Path) -> None:
    doc = Document()
    # Configure Heading 1 style
    h1 = doc.styles["Heading 1"]
    h1.font.size = Pt(18)
    h1.font.color.rgb = RGBColor(0x1F, 0x4E, 0x79)
    h1.font.bold = True
    # Configure Heading 2 style
    h2 = doc.styles["Heading 2"]
    h2.font.size = Pt(14)
    h2.font.color.rgb = RGBColor(0x2E, 0x74, 0xB5)
    h2.font.bold = True
    # Clear default empty paragraph — doc will be blank
    for para in list(doc.paragraphs):
        para._element.getparent().remove(para._element)
    doc.save(path)


def create_pptx_template(path: Path) -> None:
    prs = Presentation()
    # No slides — the template only carries the slide master / theme.
    # Add slides via generation scripts; they inherit the master theme.
    prs.save(path)


def main() -> None:
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

    targets = [
        (TEMPLATES_DIR / "recap.xlsx",        create_xlsx_template),
        (TEMPLATES_DIR / "report.docx",       create_docx_template),
        (TEMPLATES_DIR / "presentation.pptx", create_pptx_template),
    ]

    for path, creator in targets:
        if path.exists():
            print(f"skip  {path.relative_to(Path.cwd())}  (already exists — delete to recreate)")
        else:
            creator(path)
            print(f"ok    {path.relative_to(Path.cwd())}")

    print("\nOpen templates/ in LibreOffice / Word / PowerPoint to apply branding.")
    print("Generation scripts will load them automatically on next run.")


if __name__ == "__main__":
    main()
