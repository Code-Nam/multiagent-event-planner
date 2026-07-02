"""
generate_xlsx.py — AGEVP Phase 2

Reads a JSON spec file (xlsx_recap shape) and produces an .xlsx workbook
with one worksheet per sheet entry, bold header row, and auto-fitted
column widths. Output filename is derived from the input spec filename stem.

The branded template (Modele_AGEVP_Suivi.xlsx) is duplicated as a style base:
the title band above the header (title + subtitle rows, merged across the
table width) is replicated with the spec's title/event/date, and the header
row's font/fill/alignment are lifted onto each generated sheet's header.
Because a recap holds arbitrarily many sheets with varying columns, this
format stays table-driven rather than {{TOKEN}}-driven. Pass --template to
override the template path explicitly.

Usage:
    python scripts/generate_xlsx.py <json-spec-path> [--template <path>] [--output <dir>]
"""

import sys
import argparse
from copy import copy as _copy
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Missing dependency. Run: pip install -r scripts/requirements.txt", file=sys.stderr)
    sys.exit(1)

from token_fill import load_spec

TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
DEFAULT_TEMPLATE = TEMPLATES_DIR / "Modele_AGEVP_Suivi.xlsx"


def auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet, min_row: int = 1) -> None:
    """Set column widths based on the maximum cell content length.

    min_row skips the merged title band so its long text does not blow up
    the first column's width.
    """
    for col in ws.iter_cols(min_row=min_row):
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except (TypeError, AttributeError):
                continue  # MergedCell and formula cells may not expose .value
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _extract_template_styles(template: Path):
    """Return (header_style, title_band) from the template's first sheet.

    header_style: (font, fill, alignment) of the first bold+filled row.
    title_band:   [(font, alignment, row_height), ...] for the rows above it —
                  the branded title/subtitle band replicated on each sheet.
    """
    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]
    band = []
    for row in ws.iter_rows():
        c = row[0]
        try:
            has_fill = (
                c.fill
                and c.fill.fgColor
                and c.fill.fgColor.rgb not in ("00000000", "FFFFFFFF")
            )
        except (AttributeError, TypeError):
            print(
                "WARNING: could not read header style from template — using plain bold",
                file=sys.stderr,
            )
            has_fill = False
        if has_fill and c.font and c.font.bold:
            return (_copy(c.font), _copy(c.fill), _copy(c.alignment)), band
        band.append((_copy(c.font), _copy(c.alignment), ws.row_dimensions[c.row].height))
    return None, []


def _base_workbook(template: Path | None):
    """Return (workbook, header_style, title_band); see _extract_template_styles."""
    if template and template.exists():
        header_style, title_band = _extract_template_styles(template)
        wb = openpyxl.load_workbook(template)
        for name in list(wb.sheetnames):
            del wb[name]
        return wb, header_style, title_band
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb, None, []


def generate_xlsx(spec_path: Path, output_dir: Path, template: Path | None = None) -> Path:
    """
    Build an .xlsx workbook from an xlsx_recap JSON spec.

    Args:
        spec_path:  Path to the JSON spec file.
        output_dir: Directory where the .xlsx file will be written.
        template:   Optional path to an .xlsx template for branding.
                    Defaults to templates/recap.xlsx when present.

    Returns:
        Absolute path of the created .xlsx file.

    Raises:
        SystemExit: on any I/O or validation error.
    """
    spec = load_spec(spec_path, "xlsx_recap")
    sheets: list = spec.get("sheets", [])

    if not sheets:
        print("ERROR: spec contains no sheets", file=sys.stderr)
        sys.exit(1)

    resolved_template = template if template is not None else DEFAULT_TEMPLATE
    wb, header_style, title_band = _base_workbook(resolved_template)

    title: str = spec.get("title", "")
    subtitle = " — ".join(v for v in (spec.get("event", ""), spec.get("date", "")) if v)
    band_texts = [v for v in (title, subtitle) if v] if title_band else []

    for sheet_spec in sheets:
        sheet_name: str = sheet_spec.get("name", "Sheet")
        headers: list = sheet_spec.get("headers", [])
        rows: list = sheet_spec.get("rows", [])

        ws = wb.create_sheet(title=sheet_name)

        # Branded title band above the header, merged across the table width.
        n_cols = max(len(headers), 1)
        for r, text in enumerate(band_texts, 1):
            cell = ws.cell(row=r, column=1, value=text)
            b_font, b_align, b_height = title_band[min(r - 1, len(title_band) - 1)]
            cell.font = _copy(b_font)
            cell.alignment = _copy(b_align)
            if b_height:
                ws.row_dimensions[r].height = b_height
            if n_cols > 1:
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)

        if headers:
            ws.append(headers)
            header_row = len(band_texts) + 1
            h_font, h_fill, h_align = header_style if header_style else (None, None, None)
            for cell in ws[header_row]:
                if h_font:
                    cell.font = _copy(h_font)
                else:
                    cell.font = Font(bold=True)
                if h_fill:
                    cell.fill = _copy(h_fill)
                if h_align:
                    cell.alignment = _copy(h_align)

        for row in rows:
            ws.append(row)

        auto_fit_columns(ws, min_row=len(band_texts) + 1)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{spec_path.stem}.xlsx"

    try:
        wb.save(out_path)
    except OSError as exc:
        print(f"ERROR: cannot write {out_path}: {exc}", file=sys.stderr)
        sys.exit(1)

    return out_path.resolve()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate an .xlsx workbook from an xlsx_recap JSON spec."
    )
    parser.add_argument("spec", help="Path to the JSON spec file")
    parser.add_argument(
        "--template",
        default=None,
        help=f"Path to an .xlsx template for branding (default: {DEFAULT_TEMPLATE})",
    )
    parser.add_argument(
        "--output",
        default="output",
        help="Output directory (default: output/)",
    )
    args = parser.parse_args()

    result = generate_xlsx(
        spec_path=Path(args.spec),
        output_dir=Path(args.output),
        template=Path(args.template) if args.template else None,
    )
    print(result)
    sys.exit(0)
